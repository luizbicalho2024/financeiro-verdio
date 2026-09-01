import sys
import os
import re
import io
import zipfile
import unicodedata
import calendar
import hashlib
from datetime import datetime
from typing import Dict, List, Tuple, Optional

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app_core.ui import apply_branding, render_sidebar
from app_core.spreadsheet_reader import read_raw_spreadsheet
from app_core.inventory_import import SUPPORTED_TYPES

import streamlit as st
import pandas as pd
import numpy as np
from fpdf import FPDF
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import user_management_db as umdb
from mongo_config import db

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

REQUIRED_COLUMNS = [
    "Cliente", "Terminal", "Data Ativação", "Data Desativação", "Dias Ativos Mês",
    "Suspenso Dias Mes", "Nº Equipamento", "Condição"
]

COLUMN_ALIASES = {
    "equipamento": "Nº Equipamento",
    "n equipamento": "Nº Equipamento",
    "nº equipamento": "Nº Equipamento",
    "numero equipamento": "Nº Equipamento",
    "número equipamento": "Nº Equipamento",
    "suspenso dias mes": "Suspenso Dias Mes",
    "suspenso dias mês": "Suspenso Dias Mes",
    "dias ativos mes": "Dias Ativos Mês",
    "dias ativos mês": "Dias Ativos Mês",
    "data ativacao": "Data Ativação",
    "data ativação": "Data Ativação",
    "data desativacao": "Data Desativação",
    "data desativação": "Data Desativação",
    "condicao": "Condição",
    "condição": "Condição",
    "cliente": "Cliente",
    "terminal": "Terminal",
    "placa": "Placa",
    "modelo": "Modelo",
}


def _strip_accents(value: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch))


def _canonical_key(value) -> str:
    value = "" if pd.isna(value) else str(value)
    value = value.strip().replace("\n", " ").replace("\r", " ")
    value = re.sub(r"\s+", " ", value)
    value = value.replace("º", "o").replace("ª", "a")
    return _strip_accents(value).lower()


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    seen = set()
    for col in df.columns:
        raw = "" if pd.isna(col) else str(col).strip()
        key = _canonical_key(raw)
        canonical = COLUMN_ALIASES.get(key, raw)
        if canonical.lower().startswith("unnamed") or canonical == "nan" or canonical == "":
            canonical = f"Coluna_{len(seen) + 1}"
        base = canonical
        counter = 2
        while canonical in seen:
            canonical = f"{base}_{counter}"
            counter += 1
        rename[col] = canonical
        seen.add(canonical)
    return df.rename(columns=rename)


def _normalize_equipment(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    return text


def _normalize_tipo(value) -> str:
    if pd.isna(value):
        return ""
    text = _strip_accents(str(value).strip().upper())
    text = re.sub(r"\s+", " ", text)
    if text in {"SATELITAL", "SATELLITE", "SATELITE"}:
        return "SATELITE"
    return text


def _money_br(value) -> str:
    try:
        value = float(value)
    except Exception:
        value = 0.0
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _safe_float(value, default=0.0) -> float:
    if value is None or pd.isna(value):
        return float(default)
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).strip()
    if not text:
        return float(default)
    text = text.replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return float(default)


def sanitize_id(name: str) -> str:
    return str(name).strip().replace("/", "-")


def _read_raw_report(file_bytes: bytes, file_name: str = "") -> pd.DataFrame:
    return read_raw_spreadsheet(file_bytes, file_name)


def _find_header_row(df_raw: pd.DataFrame) -> Optional[int]:
    max_scan = min(60, len(df_raw))
    for idx in range(max_scan):
        values = [_canonical_key(v) for v in df_raw.iloc[idx].tolist()]
        has_cliente = "cliente" in values
        has_terminal = "terminal" in values
        has_equip = any(v in {"equipamento", "n equipamento", "numero equipamento", "número equipamento"} for v in values)
        if has_cliente and has_terminal and has_equip:
            return idx
    for idx in range(max_scan):
        row_text = " ".join(str(v) for v in df_raw.iloc[idx].tolist())
        key = _canonical_key(row_text)
        if "cliente" in key and "terminal" in key:
            return idx
    return None


def _extract_report_date(df_raw: pd.DataFrame) -> pd.Timestamp:
    max_scan = min(40, len(df_raw))
    date_regex = r"(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})"

    for idx in range(max_scan):
        row_text = " ".join(str(v) for v in df_raw.iloc[idx].tolist() if not pd.isna(v))
        match = re.search(r"Data\s*Final\s*[:\-]?\s*" + date_regex, row_text, flags=re.IGNORECASE)
        if match:
            return pd.to_datetime(match.group(1).replace("-", "/"), dayfirst=True, errors="coerce")

    for idx in range(max_scan):
        row_text = " ".join(str(v) for v in df_raw.iloc[idx].tolist() if not pd.isna(v))
        dates = re.findall(date_regex, row_text)
        if dates:
            # Quando a célula contém data inicial e final, o período correto para faturamento é a última data.
            parsed = pd.to_datetime(dates[-1].replace("-", "/"), dayfirst=True, errors="coerce")
            if pd.notna(parsed):
                return parsed
    return pd.NaT


def _prepare_report_dataframe(file_bytes: bytes, file_name: str = "") -> Tuple[pd.DataFrame, pd.Timestamp]:
    df_raw = _read_raw_report(file_bytes, file_name)
    report_date = _extract_report_date(df_raw)
    header_row_idx = _find_header_row(df_raw)
    if header_row_idx is None:
        raise ValueError("Não foi possível encontrar o cabeçalho do relatório. O arquivo precisa conter as colunas Cliente, Terminal e Equipamento/Nº Equipamento.")

    df = df_raw.iloc[header_row_idx + 1:].copy()
    df.columns = [str(c).strip() for c in df_raw.iloc[header_row_idx].tolist()]
    df = _normalize_columns(df)

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError("Colunas obrigatórias ausentes no relatório: " + ", ".join(missing))

    df = df.dropna(subset=["Terminal"])
    df = df[df["Terminal"].astype(str).str.strip() != ""]
    df = df[df["Cliente"].astype(str).str.strip().str.lower() != "cliente"]
    df["Cliente"] = df["Cliente"].astype(str).str.strip()
    df["Terminal"] = df["Terminal"].astype(str).str.strip()
    df["Nº Equipamento"] = df["Nº Equipamento"].apply(_normalize_equipment)

    for col in ["Data Ativação", "Data Desativação"]:
        df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
    for col in ["Dias Ativos Mês", "Suspenso Dias Mes"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    if pd.isna(report_date):
        dates = pd.concat([df["Data Ativação"].dropna(), df["Data Desativação"].dropna()])
        report_date = dates.max() if not dates.empty else pd.Timestamp(datetime.now())

    return df, report_date


def _prepare_inventory(tracker_inventory: List[dict]) -> pd.DataFrame:
    df_inventory = pd.DataFrame(tracker_inventory or [])
    if df_inventory.empty:
        raise ValueError("Nenhum dado de estoque de rastreadores encontrado.")
    df_inventory = _normalize_columns(df_inventory)
    if "Nº Equipamento" not in df_inventory.columns:
        raise ValueError("O estoque de rastreadores não possui a coluna 'Nº Equipamento'.")
    if "Tipo" not in df_inventory.columns:
        raise ValueError("O estoque de rastreadores não possui a coluna 'Tipo'.")
    if "Modelo" not in df_inventory.columns:
        df_inventory["Modelo"] = ""
    df_inventory["Nº Equipamento"] = df_inventory["Nº Equipamento"].apply(_normalize_equipment)
    df_inventory["Tipo"] = df_inventory["Tipo"].apply(_normalize_tipo)
    df_inventory = df_inventory.drop_duplicates(subset=["Nº Equipamento"], keep="last")
    return df_inventory


def _calculate_billing(df: pd.DataFrame, df_inventory: pd.DataFrame, report_date: pd.Timestamp, prices_by_type: Dict[str, float]) -> Tuple[str, pd.DataFrame, List[str]]:
    report_month = int(report_date.month)
    report_year = int(report_date.year)
    dias_no_mes = calendar.monthrange(report_year, report_month)[1]
    month_start = pd.Timestamp(year=report_year, month=report_month, day=1)
    month_end = pd.Timestamp(year=report_year, month=report_month, day=dias_no_mes)
    periodo_relatorio = f"{MESES_PT[report_month]} de {report_year}"

    df_merged = pd.merge(df, df_inventory[["Nº Equipamento", "Modelo", "Tipo"]], on="Nº Equipamento", how="left", suffixes=("", "_Estoque"))
    if "Modelo_Estoque" in df_merged.columns:
        df_merged["Modelo"] = df_merged["Modelo"].where(df_merged["Modelo"].notna() & (df_merged["Modelo"].astype(str).str.strip() != ""), df_merged["Modelo_Estoque"])
        df_merged = df_merged.drop(columns=["Modelo_Estoque"])

    df_merged["Tipo"] = df_merged["Tipo"].apply(_normalize_tipo)

    inventory_serials = set(
        df_inventory["Nº Equipamento"].dropna().astype(str).str.strip().tolist()
    )
    missing_inventory_mask = ~df_merged["Nº Equipamento"].astype(str).str.strip().isin(
        inventory_serials
    )

    model_type_source = df_inventory[["Modelo", "Tipo"]].copy()
    model_type_source["_model_key"] = model_type_source["Modelo"].apply(_canonical_key)
    model_type_source["Tipo"] = model_type_source["Tipo"].apply(_normalize_tipo)
    model_type_source = model_type_source[
        model_type_source["_model_key"].ne("")
        & model_type_source["Tipo"].ne("")
    ]

    model_type_map = {}
    if not model_type_source.empty:
        grouped_types = model_type_source.groupby("_model_key")["Tipo"].agg(
            lambda values: sorted(set(value for value in values if value))
        )
        model_type_map = {
            model_key: types[0]
            for model_key, types in grouped_types.items()
            if len(types) == 1
        }

    missing_type_mask = df_merged["Tipo"].eq("")
    inferred_types = df_merged["Modelo"].apply(_canonical_key).map(model_type_map).fillna("")
    df_merged.loc[missing_type_mask, "Tipo"] = inferred_types[missing_type_mask]

    not_found = sorted(
        [
            x
            for x in df_merged.loc[
                missing_inventory_mask,
                "Nº Equipamento",
            ]
            .dropna()
            .unique()
            .tolist()
            if str(x).strip()
        ]
    )

    normalized_prices = {_normalize_tipo(k): _safe_float(v) for k, v in (prices_by_type or {}).items()}
    df_merged["Valor Unitario"] = df_merged["Tipo"].map(normalized_prices).fillna(0.0).astype(float)

    ativacao = df_merged["Data Ativação"]
    desativacao = df_merged["Data Desativação"]
    condicao = df_merged["Condição"].astype(str).map(lambda x: _strip_accents(x).strip().lower())

    ativado_no_mes = ativacao.notna() & (ativacao.dt.month == report_month) & (ativacao.dt.year == report_year)
    desativado_no_mes = desativacao.notna() & (desativacao.dt.month == report_month) & (desativacao.dt.year == report_year)
    suspenso = condicao.str.contains("suspenso", na=False) | (df_merged["Suspenso Dias Mes"] > 0)

    df_merged["Categoria"] = np.select(
        [ativado_no_mes & desativado_no_mes, desativado_no_mes, ativado_no_mes, suspenso],
        ["Ativado e Desativado no Mês", "Desativado", "Ativado no Mês", "Suspenso"],
        default="Cheio",
    )

    active_start = ativacao.where(ativacao.notna() & (ativacao > month_start), month_start)
    active_end = desativacao.where(desativacao.notna() & (desativacao < month_end), month_end)

    # Casos fora do mês não devem faturar dias negativos.
    active_days = (active_end - active_start).dt.days + 1
    active_days = active_days.where(ativacao.isna() | (ativacao <= month_end), 0)
    active_days = active_days.where(desativacao.isna() | (desativacao >= month_start), 0)
    active_days = pd.to_numeric(active_days, errors="coerce").fillna(0).clip(lower=0, upper=dias_no_mes)

    suspended_days = pd.to_numeric(df_merged["Suspenso Dias Mes"], errors="coerce").fillna(0).clip(lower=0, upper=dias_no_mes)
    df_merged["Dias Ativos Calculado"] = active_days.astype(int)
    df_merged["Dias a Faturar"] = (active_days - suspended_days).clip(lower=0, upper=dias_no_mes).round(0).astype(int)
    df_merged["Valor a Faturar"] = ((df_merged["Valor Unitario"] / dias_no_mes) * df_merged["Dias a Faturar"]).round(2)

    # Itens sem tipo/preço devem aparecer para conferência, mas com faturamento zerado para evitar cobrança indevida.
    df_merged.loc[df_merged["Tipo"].eq(""), "Valor a Faturar"] = 0.0

    return periodo_relatorio, df_merged, not_found


def _format_excel_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if cell.column_letter:
                header = ws.cell(row=1, column=cell.column).value
                if header and "Valor" in str(header):
                    cell.number_format = 'R$ #,##0.00'
                if header and "Data" in str(header):
                    cell.number_format = 'DD/MM/YYYY'
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 45))


def _clean_export_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["Faturar", "Coluna_1", "Coluna_2", "Coluna_3"]:
        if col in df.columns:
            df = df.drop(columns=[col])
    preferred = [
        "Cliente", "Terminal", "Nº Equipamento", "Placa", "Modelo", "Tipo", "Condição", "Categoria",
        "Data Ativação", "Data Desativação", "Dias Ativos Mês", "Dias Ativos Calculado",
        "Suspenso Dias Mes", "Dias a Faturar", "Valor Unitario", "Valor a Faturar",
    ]
    cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
    return df[cols]


def _to_excel_named_sheets(sheets: Dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = name[:31]
            _clean_export_df(df).to_excel(writer, index=False, sheet_name=safe_name)
        for ws in writer.book.worksheets:
            _format_excel_sheet(ws)
    return output.getvalue()


def _pdf_bytes(pdf: FPDF) -> bytes:
    data = pdf.output(dest="S")
    if isinstance(data, (bytes, bytearray)):
        return bytes(data)
    return str(data).encode("latin-1", errors="replace")


class PDF(FPDF):
    def header(self):
        try:
            page_width = self.w - self.l_margin - self.r_margin
            self.image("imgs/header1.png", x=self.l_margin, y=8, w=page_width)
        except Exception:
            self.set_font("Arial", "B", 20)
            self.cell(0, 10, "Uzzipay Soluções", 0, 1, "L")
            self.ln(15)

    def footer(self):
        try:
            self.set_y(-35)
            page_width = self.w - self.l_margin - self.r_margin
            self.image("imgs/footer1.png", x=self.l_margin, y=self.get_y(), w=page_width)
        except Exception:
            self.set_y(-15)
            self.set_font("Arial", "I", 8)
            self.cell(0, 10, f"Página {self.page_no()}", 0, 0, "C")


def create_pdf_report(nome_cliente, periodo, totais, df_cheio, df_ativados, df_desativados, df_suspensos, df_ativados_desativados=None):
    df_ativados_desativados = df_ativados_desativados if df_ativados_desativados is not None else pd.DataFrame()
    pdf = PDF(orientation="L")
    pdf.set_margins(10, 40, 10)
    pdf.set_top_margin(40)
    pdf.set_auto_page_break(auto=True, margin=35)
    pdf.add_page()

    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Resumo do Faturamento", 0, 1, "C")
    pdf.ln(4)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 7, f"Cliente: {nome_cliente}", 0, 1, "L")
    pdf.cell(0, 7, f"Período: {periodo}", 0, 1, "L")
    pdf.ln(4)

    table_width = pdf.w - pdf.l_margin - pdf.r_margin
    col_width = table_width / 5
    pdf.set_font("Arial", "B", 9)
    for title in ["Nº Fat. Cheio", "Nº Fat. Proporcional", "Nº Suspensos", "Total GPRS", "Total Satelitais"]:
        pdf.cell(col_width, 8, title, 1, 0, "C")
    pdf.ln()
    pdf.set_font("Arial", "", 9)
    for value in [totais.get("terminais_cheio", 0), totais.get("terminais_proporcional", 0), totais.get("terminais_suspensos", 0), totais.get("terminais_gprs", 0), totais.get("terminais_satelitais", 0)]:
        pdf.cell(col_width, 8, str(value), 1, 0, "C")
    pdf.ln(10)

    pdf.set_font("Arial", "B", 11)
    pdf.cell(table_width / 2, 8, "Faturamento Cheio", 1, 0, "C")
    pdf.cell(table_width / 2, 8, "Faturamento Proporcional", 1, 1, "C")
    pdf.set_font("Arial", "", 11)
    pdf.cell(table_width / 2, 8, _money_br(totais.get("cheio", 0)), 1, 0, "C")
    pdf.cell(table_width / 2, 8, _money_br(totais.get("proporcional", 0)), 1, 1, "C")
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"FATURAMENTO TOTAL: {_money_br(totais.get('geral', 0))}", 1, 1, "C")
    pdf.ln(6)

    def draw_table(title, df, col_widths, available_cols):
        if df is None or df.empty:
            return
        if pdf.get_y() > pdf.h - 65:
            pdf.add_page()
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 8, title, 0, 1, "L")
        cols = [c for c in available_cols if c in df.columns]
        pdf.set_font("Arial", "B", 7)
        for col in cols:
            pdf.cell(col_widths.get(col, 20), 7, col, 1, 0, "C")
        pdf.ln()
        pdf.set_font("Arial", "", 6)
        for _, row in df.iterrows():
            if pdf.get_y() > pdf.h - 42:
                pdf.add_page()
                pdf.set_font("Arial", "B", 7)
                for col in cols:
                    pdf.cell(col_widths.get(col, 20), 7, col, 1, 0, "C")
                pdf.ln()
                pdf.set_font("Arial", "", 6)
            for col in cols:
                value = row.get(col, "")
                if pd.isna(value):
                    text = ""
                elif "Data" in col:
                    text = pd.to_datetime(value).strftime("%d/%m/%Y") if pd.notna(pd.to_datetime(value, errors="coerce")) else ""
                elif "Valor" in col:
                    text = _money_br(value)
                else:
                    text = str(value)[:38]
                pdf.cell(col_widths.get(col, 20), 6, text, 1, 0, "C")
            pdf.ln()
        pdf.ln(4)

    widths_cheio = {"Terminal": 35, "Nº Equipamento": 34, "Placa": 25, "Modelo": 45, "Tipo": 22, "Dias a Faturar": 24, "Valor Unitario": 28, "Valor a Faturar": 30}
    cols_cheio = list(widths_cheio.keys())
    widths_prop = {"Terminal": 24, "Nº Equipamento": 26, "Modelo": 32, "Tipo": 18, "Data Ativação": 24, "Data Desativação": 24, "Dias Ativos Calculado": 25, "Suspenso Dias Mes": 24, "Dias a Faturar": 22, "Valor Unitario": 26, "Valor a Faturar": 28}
    cols_prop = list(widths_prop.keys())
    draw_table("Detalhamento do Faturamento Cheio", df_cheio, widths_cheio, cols_cheio)
    draw_table("Detalhamento Proporcional - Ativações no Mês", df_ativados, widths_prop, cols_prop)
    draw_table("Detalhamento Proporcional - Desativações no Mês", df_desativados, widths_prop, cols_prop)
    draw_table("Detalhamento Proporcional - Ativados e Desativados no Mês", df_ativados_desativados, widths_prop, cols_prop)
    draw_table("Detalhamento dos Terminais Suspensos", df_suspensos, widths_prop, cols_prop)

    return _pdf_bytes(pdf)


def split_categories(df_aprovado: pd.DataFrame):
    df_cheio = df_aprovado[df_aprovado["Categoria"] == "Cheio"].copy()
    df_ativados = df_aprovado[df_aprovado["Categoria"] == "Ativado no Mês"].copy()
    df_desativados = df_aprovado[df_aprovado["Categoria"] == "Desativado"].copy()
    df_ativados_desativados = df_aprovado[df_aprovado["Categoria"] == "Ativado e Desativado no Mês"].copy()
    df_suspensos = df_aprovado[df_aprovado["Categoria"] == "Suspenso"].copy()
    return df_cheio, df_ativados, df_desativados, df_suspensos, df_ativados_desativados


def proportional_categories() -> List[str]:
    return ["Ativado no Mês", "Desativado", "Ativado e Desativado no Mês", "Suspenso"]


def build_totals(df_aprovado: pd.DataFrame) -> dict:
    df_cheio, df_ativados, df_desativados, df_suspensos, df_ativados_desativados = split_categories(df_aprovado)
    total_cheio = float(df_cheio["Valor a Faturar"].sum())
    total_prop = float(df_aprovado[df_aprovado["Categoria"].isin(proportional_categories())]["Valor a Faturar"].sum())
    return {
        "cheio": total_cheio,
        "proporcional": total_prop,
        "geral": total_cheio + total_prop,
        "terminais_cheio": int(len(df_cheio)),
        "terminais_proporcional": int(len(df_aprovado[df_aprovado["Categoria"].isin(["Ativado no Mês", "Desativado", "Ativado e Desativado no Mês"])])),
        "terminais_suspensos": int(len(df_suspensos)),
        "terminais_gprs": int(len(df_aprovado[df_aprovado["Tipo"] == "GPRS"])),
        "terminais_satelitais": int(len(df_aprovado[df_aprovado["Tipo"] == "SATELITE"])),
    }

# --- 1. CONFIGURAÇÃO E AUTENTICAÇÃO ---
st.set_page_config(layout="wide", page_title="Faturamento em Lote", page_icon="imgs/v-c.png")
apply_branding()
if "user_info" not in st.session_state:
    st.error("🔒 Acesso Negado! Por favor, faça login para visualizar esta página.")
    st.stop()

render_sidebar()


def _load_contract_prices() -> Dict[str, Dict[str, float]]:
    docs = db.collection("client_contracts").stream()
    contracts = {doc.id: doc.to_dict() for doc in docs}
    result = {}
    for doc_id, data in contracts.items():
        client_name = str(data.get("cliente", doc_id)).strip()
        prices = data.get("precos_por_tipo", {}) or {}
        normalized = {_normalize_tipo(k): _safe_float(v) for k, v in prices.items()}
        result[client_name] = normalized
        result[sanitize_id(client_name)] = normalized
    return result


@st.cache_data(show_spinner=False)
def processar_planilha_lote(file_bytes, file_name, tracker_inventory):
    try:
        df, report_date = _prepare_report_dataframe(file_bytes, file_name)
        df_inventory = _prepare_inventory(tracker_inventory)
        contracts = _load_contract_prices()

        # Calcula cliente por cliente para respeitar preço contratado individual.
        frames = []
        not_found_all = set()
        periodo_relatorio = None
        for cliente, df_cliente in df.groupby("Cliente", dropna=False):
            cliente = str(cliente).strip()
            prices = contracts.get(cliente) or contracts.get(sanitize_id(cliente)) or {}
            periodo_relatorio, df_calc, not_found = _calculate_billing(df_cliente.copy(), df_inventory, report_date, prices)
            frames.append(df_calc)
            not_found_all.update(not_found)

        df_final = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        return periodo_relatorio, df_final, sorted(not_found_all), None
    except Exception as e:
        return None, None, [], f"Ocorreu um erro ao processar o lote: {e}"


def generate_master_excel(df_aprovado):
    resumo_data = []
    for cliente, df_cliente in df_aprovado.groupby("Cliente"):
        detalhes_modelos = []
        for tipo, df_tipo in df_cliente.groupby("Tipo"):
            val = _safe_float(df_tipo["Valor Unitario"].max())
            qtd = len(df_tipo)
            detalhes_modelos.append(f"{tipo or 'SEM TIPO'}: {qtd} un. a {_money_br(val)}")
        totais = build_totals(df_cliente)
        resumo_data.append({
            "Cliente": cliente,
            "Detalhes Contratos (Modelos)": " | ".join(detalhes_modelos),
            "Qtd Terminais Faturados": len(df_cliente),
            "Qtd Cheio": totais["terminais_cheio"],
            "Qtd Proporcional": totais["terminais_proporcional"],
            "Qtd Suspensos": totais["terminais_suspensos"],
            "Valor Total a Faturar": totais["geral"],
        })
    df_resumo = pd.DataFrame(resumo_data).sort_values("Cliente") if resumo_data else pd.DataFrame()
    return _to_excel_named_sheets({
        "Resumo Faturamento Lote": df_resumo,
        "Todos os Terminais": df_aprovado,
    })


def create_zip_of_pdfs(df_aprovado, periodo_relatorio):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for cliente in sorted(df_aprovado["Cliente"].dropna().unique()):
            df_cliente = df_aprovado[df_aprovado["Cliente"] == cliente].copy()
            df_cheio, df_ativados, df_desativados, df_suspensos, df_ativados_desativados = split_categories(df_cliente)
            totais = build_totals(df_cliente)
            pdf_bytes = create_pdf_report(cliente, periodo_relatorio, totais, df_cheio, df_ativados, df_desativados, df_suspensos, df_ativados_desativados)
            safe_cliente = re.sub(r"[^A-Za-z0-9]+", "_", str(cliente)).strip("_") or "Cliente"
            zip_file.writestr(f"Faturamento_{safe_cliente}.pdf", pdf_bytes)
    return zip_buffer.getvalue()


def salvar_historico_lote(df_aprovado, periodo_relatorio):
    success_count = 0
    failed_clients = []
    cols_to_save = [
        "Terminal", "Nº Equipamento", "Placa", "Frota", "Modelo", "Tipo", "Condição", "Categoria",
        "Data Ativação", "Data Desativação", "Dias Ativos Mês", "Dias Ativos Calculado",
        "Suspenso Dias Mes", "Dias a Faturar", "Valor Unitario", "Valor a Faturar",
    ]

    for cliente in sorted(df_aprovado["Cliente"].dropna().unique()):
        df_cliente = df_aprovado[df_aprovado["Cliente"] == cliente].copy()
        totais = build_totals(df_cliente)
        log_data = {
            "cliente": cliente,
            "periodo_relatorio": periodo_relatorio,
            "valor_total": totais["geral"],
            "terminais_cheio": totais["terminais_cheio"],
            "terminais_proporcional": totais["terminais_proporcional"],
            "terminais_suspensos": totais["terminais_suspensos"],
            "terminais_gprs": totais["terminais_gprs"],
            "terminais_satelitais": totais["terminais_satelitais"],
            "valor_unitario_gprs": _safe_float(df_cliente[df_cliente["Tipo"] == "GPRS"]["Valor Unitario"].max()) if not df_cliente[df_cliente["Tipo"] == "GPRS"].empty else 0.0,
            "valor_unitario_satelital": _safe_float(df_cliente[df_cliente["Tipo"] == "SATELITE"]["Valor Unitario"].max()) if not df_cliente[df_cliente["Tipo"] == "SATELITE"].empty else 0.0,
        }
        clean = _clean_export_df(df_cliente)
        detalhes_itens = clean[[column for column in cols_to_save if column in clean.columns]].to_dict(orient="records")
        if umdb.log_faturamento(log_data, detalhes_itens, notify=False):
            success_count += 1
        else:
            failed_clients.append(str(cliente))

    if failed_clients:
        st.session_state["lote_salvo"] = False
        st.error(
            "O lote não foi fechado porque alguns clientes falharam ao salvar: "
            + ", ".join(failed_clients[:20])
            + ("..." if len(failed_clients) > 20 else "")
        )
        return False

    closed = umdb.close_billing_month(
        periodo_relatorio,
        total_clientes=int(df_aprovado["Cliente"].nunique()),
        total_terminais=int(len(df_aprovado)),
        faturamento_total=float(df_aprovado["Valor a Faturar"].sum()),
    )
    st.session_state["lote_salvo"] = bool(closed)
    if closed:
        st.success(f"{success_count} clientes salvos e mês fechado para análise comercial.")
    return bool(closed)

# --- 4. INTERFACE ---
st.subheader("FINANCEIRO - Processamento de Faturamento em Lote Verdio")
st.info(
    "Envie uma ou várias planilhas. Cada arquivo é processado de forma independente, "
    "o período é identificado automaticamente e os resultados são agrupados por mês. "
    "Isso permite carregar vários meses históricos de uma única vez."
)

uploaded_files = st.file_uploader(
    "Selecione um ou mais relatórios de faturamento",
    type=["xls", "xlsx", "csv"],
    accept_multiple_files=True,
    key="billing_multi_files",
)

st.markdown("---")


def _period_sort_key(period_label: str):
    normalized_months = {
        _strip_accents(name).lower(): month
        for month, name in MESES_PT.items()
    }
    match = re.match(
        r"^\s*([A-Za-zÀ-ÿ]+)\s+de\s+(\d{4})\s*$",
        str(period_label or ""),
        flags=re.IGNORECASE,
    )
    if not match:
        return (9999, 99, str(period_label or ""))

    month_name = _strip_accents(match.group(1)).lower()
    return (
        int(match.group(2)),
        normalized_months.get(month_name, 99),
        str(period_label or ""),
    )


def _safe_widget_key(value: str) -> str:
    return re.sub(
        r"[^A-Za-z0-9_]+",
        "_",
        str(value or ""),
    ).strip("_")[:100]


def _billing_batch_signature(files, tracker_inventory) -> str:
    digest = hashlib.sha256()

    for uploaded in files:
        payload = uploaded.getvalue()
        digest.update(str(uploaded.name or "").encode("utf-8", errors="ignore"))
        digest.update(str(len(payload)).encode("ascii"))
        digest.update(hashlib.sha256(payload).digest())

    inventory_rows = sorted(
        (
            _normalize_equipment(item.get("Nº Equipamento", "")),
            str(item.get("Modelo", "") or "").strip(),
            _normalize_tipo(item.get("Tipo", "")),
        )
        for item in (tracker_inventory or [])
    )
    for serial_number, model, equipment_type in inventory_rows:
        digest.update(
            f"{serial_number}|{model}|{equipment_type}\n".encode(
                "utf-8",
                errors="ignore",
            )
        )

    return digest.hexdigest()


def _frame_signature(df: pd.DataFrame) -> str:
    clean = _clean_export_df(df).copy()
    normalized = clean.fillna("").astype(str)
    hashes = pd.util.hash_pandas_object(normalized, index=True)
    digest = hashlib.sha1()
    digest.update("|".join(normalized.columns.astype(str)).encode("utf-8"))
    digest.update(hashes.values.tobytes())
    return digest.hexdigest()


def _build_missing_inventory_candidates(
    processed_periods,
    tracker_inventory,
) -> pd.DataFrame:
    existing = {
        _normalize_equipment(item.get("Nº Equipamento", ""))
        for item in (tracker_inventory or [])
    }
    existing.discard("")

    frames = []
    for periodo, info in (processed_periods or {}).items():
        frame = info.get("df")
        if frame is None or frame.empty or "Nº Equipamento" not in frame.columns:
            continue

        candidate = frame.copy()
        candidate["Nº Equipamento"] = candidate["Nº Equipamento"].apply(
            _normalize_equipment
        )
        candidate = candidate[
            candidate["Nº Equipamento"].ne("")
            & ~candidate["Nº Equipamento"].isin(existing)
        ].copy()

        if candidate.empty:
            continue

        if "Modelo" not in candidate.columns:
            candidate["Modelo"] = ""
        if "Tipo" not in candidate.columns:
            candidate["Tipo"] = ""

        candidate["Modelo"] = (
            candidate["Modelo"].fillna("").astype(str).str.strip()
        )
        candidate["Tipo"] = candidate["Tipo"].apply(_normalize_tipo)
        candidate["Período detectado"] = periodo

        frames.append(
            candidate[
                [
                    "Nº Equipamento",
                    "Modelo",
                    "Tipo",
                    "Período detectado",
                ]
            ]
        )

    if not frames:
        return pd.DataFrame(
            columns=[
                "Nº Equipamento",
                "Modelo",
                "Tipo",
                "Período detectado",
            ]
        )

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(
        subset=["Nº Equipamento"],
        keep="last",
    )
    return combined.sort_values(
        ["Modelo", "Nº Equipamento"],
        kind="stable",
    ).reset_index(drop=True)


def _clear_billing_batch_cache() -> None:
    for key in [
        "billing_processed_batch_cache",
        "billing_export_cache",
    ]:
        st.session_state.pop(key, None)


def processar_arquivos_historicos(files, tracker_inventory):
    grouped = {}
    errors = []

    for uploaded in files:
        periodo, frame, not_found, error = processar_planilha_lote(
            uploaded.getvalue(),
            uploaded.name,
            tracker_inventory,
        )

        if error or frame is None or frame.empty:
            errors.append(
                {
                    "Arquivo": uploaded.name,
                    "Erro": error or "Nenhum registro faturável encontrado.",
                }
            )
            continue

        bucket = grouped.setdefault(
            periodo,
            {
                "frames": [],
                "files": [],
                "not_found": set(),
            },
        )
        bucket["frames"].append(frame)
        bucket["files"].append(uploaded.name)
        bucket["not_found"].update(not_found or [])

    processed = {}

    for periodo, bucket in grouped.items():
        combined = pd.concat(
            bucket["frames"],
            ignore_index=True,
        )

        dedup_columns = [
            column
            for column in [
                "Cliente",
                "Terminal",
                "Nº Equipamento",
            ]
            if column in combined.columns
        ]

        duplicates_removed = 0
        if dedup_columns:
            before = len(combined)
            combined = combined.drop_duplicates(
                subset=dedup_columns,
                keep="last",
            ).reset_index(drop=True)
            duplicates_removed = before - len(combined)

        processed[periodo] = {
            "df": combined,
            "files": bucket["files"],
            "not_found": sorted(bucket["not_found"]),
            "duplicates_removed": int(duplicates_removed),
        }

    return processed, errors


if uploaded_files:
    tracker_inventory = umdb.get_tracker_inventory()

    if not tracker_inventory:
        st.warning(
            "O estoque está vazio. Importe o estoque antes de processar o faturamento."
        )
        st.stop()

    batch_signature = _billing_batch_signature(
        uploaded_files,
        tracker_inventory,
    )
    batch_cache = st.session_state.get("billing_processed_batch_cache")

    if (
        isinstance(batch_cache, dict)
        and batch_cache.get("signature") == batch_signature
    ):
        processed_periods = batch_cache.get("processed_periods", {})
        processing_errors = batch_cache.get("processing_errors", [])
        st.caption(
            "Lote reaproveitado da memória da sessão — as planilhas não foram "
            "reprocessadas neste clique."
        )
    else:
        with st.spinner(
            f"Processando {len(uploaded_files)} arquivo(s) e identificando os períodos..."
        ):
            processed_periods, processing_errors = processar_arquivos_historicos(
                uploaded_files,
                tracker_inventory,
            )

        st.session_state["billing_processed_batch_cache"] = {
            "signature": batch_signature,
            "processed_periods": processed_periods,
            "processing_errors": processing_errors,
        }
        st.session_state.pop("billing_export_cache", None)

    cache_action_col, cache_info_col = st.columns([1, 4])
    with cache_action_col:
        if st.button(
            "Reprocessar lote",
            key="force_reprocess_billing_batch",
            use_container_width=True,
        ):
            _clear_billing_batch_cache()
            processar_planilha_lote.clear()
            st.rerun()
    with cache_info_col:
        st.caption(
            "Use Reprocessar lote apenas quando preços, contratos ou estoque "
            "forem alterados durante esta mesma carga."
        )

    if processing_errors:
        st.error(
            f"{len(processing_errors)} arquivo(s) não puderam ser processados."
        )
        st.dataframe(
            pd.DataFrame(processing_errors),
            use_container_width=True,
            hide_index=True,
        )

    if not processed_periods:
        st.warning("Nenhum período válido foi processado.")
        st.stop()

    ordered_periods = sorted(
        processed_periods,
        key=_period_sort_key,
    )

    total_rows = sum(
        len(processed_periods[periodo]["df"])
        for periodo in ordered_periods
    )
    total_clients = sum(
        processed_periods[periodo]["df"]["Cliente"].nunique()
        for periodo in ordered_periods
    )

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Arquivos enviados", len(uploaded_files))
    k2.metric("Meses identificados", len(ordered_periods))
    k3.metric("Registros processados", total_rows)
    k4.metric("Clientes/mês", total_clients)

    st.markdown("### Resumo da carga")

    summary_rows = []
    for periodo in ordered_periods:
        info = processed_periods[periodo]
        frame = info["df"]
        totals = build_totals(frame)

        summary_rows.append(
            {
                "Período": periodo,
                "Arquivos": len(info["files"]),
                "Clientes": int(frame["Cliente"].nunique()),
                "Terminais": int(len(frame)),
                "Não encontrados": len(info["not_found"]),
                "Duplicados removidos": info["duplicates_removed"],
                "Valor calculado": float(totals["geral"]),
            }
        )

    st.dataframe(
        pd.DataFrame(summary_rows),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Valor calculado": st.column_config.NumberColumn(
                "Valor calculado",
                format="R$ %.2f",
            ),
        },
    )

    missing_candidates = _build_missing_inventory_candidates(
        processed_periods,
        tracker_inventory,
    )

    if not missing_candidates.empty:
        st.markdown("### Novos equipamentos detectados nas planilhas")
        inferred_count = int(missing_candidates["Tipo"].ne("").sum())
        unresolved_count = int(missing_candidates["Tipo"].eq("").sum())

        st.info(
            f"{len(missing_candidates)} equipamento(s) ainda não existem no estoque. "
            f"{inferred_count} tiveram o tipo inferido pelo Modelo já conhecido; "
            f"{unresolved_count} precisam de classificação manual."
        )

        missing_editor = st.data_editor(
            missing_candidates,
            hide_index=True,
            use_container_width=True,
            key=f"missing_inventory_editor_{batch_signature[:12]}",
            column_config={
                "Nº Equipamento": st.column_config.TextColumn(
                    "Nº Equipamento",
                    disabled=True,
                ),
                "Modelo": st.column_config.TextColumn(
                    "Modelo",
                    disabled=True,
                ),
                "Tipo": st.column_config.SelectboxColumn(
                    "Tipo para faturamento",
                    options=[""] + SUPPORTED_TYPES,
                    required=False,
                ),
                "Período detectado": st.column_config.TextColumn(
                    "Detectado em",
                    disabled=True,
                ),
            },
        )

        unresolved_after_edit = missing_editor[
            missing_editor["Tipo"].fillna("").astype(str).str.strip().eq("")
        ]

        if unresolved_after_edit.empty:
            if st.button(
                "Adicionar novos equipamentos ao estoque e recalcular",
                type="primary",
                key="save_missing_inventory_candidates",
                use_container_width=True,
            ):
                stock_to_save = missing_editor[
                    ["Nº Equipamento", "Modelo", "Tipo"]
                ].copy()

                with st.spinner(
                    f"Adicionando {len(stock_to_save)} equipamento(s) ao estoque..."
                ):
                    count = umdb.update_tracker_inventory(
                        stock_to_save,
                        source_file="Descoberta automática no faturamento histórico",
                    )

                if count is not None:
                    st.success(
                        f"{count} equipamento(s) adicionados/atualizados no estoque. "
                        "O lote será recalculado com os tipos e preços correspondentes."
                    )
                    _clear_billing_batch_cache()
                    processar_planilha_lote.clear()
                    st.rerun()
        else:
            st.warning(
                f"Classifique o Tipo de {len(unresolved_after_edit)} equipamento(s) "
                "antes de adicioná-los ao estoque. Não fazemos classificação automática "
                "quando o Modelo ainda não possui referência confiável."
            )

    approved_by_period = {}

    for periodo in ordered_periods:
        info = processed_periods[periodo]
        df_period = info["df"].copy()
        key_suffix = _safe_widget_key(periodo)

        with st.expander(
            f"{periodo} · {len(df_period)} terminais · "
            f"{df_period['Cliente'].nunique()} clientes",
            expanded=len(ordered_periods) == 1,
        ):
            st.caption(
                "Arquivo(s): " + ", ".join(info["files"])
            )

            if info["duplicates_removed"]:
                st.warning(
                    f"{info['duplicates_removed']} registro(s) duplicado(s) "
                    "do mesmo cliente/terminal/equipamento foram consolidados."
                )

            if info["not_found"]:
                unresolved_missing = int(
                    df_period[
                        df_period["Nº Equipamento"].isin(info["not_found"])
                        & df_period["Tipo"].fillna("").astype(str).str.strip().eq("")
                    ].shape[0]
                )
                if unresolved_missing:
                    st.warning(
                        f"{len(info['not_found'])} equipamento(s) estão ausentes do estoque; "
                        f"{unresolved_missing} ainda estão sem Tipo e permanecem com faturamento "
                        "zerado. Use a seção 'Novos equipamentos detectados nas planilhas'."
                    )
                else:
                    st.info(
                        f"{len(info['not_found'])} equipamento(s) estão ausentes do estoque, "
                        "mas tiveram o Tipo inferido pelo Modelo e já participam do cálculo. "
                        "Adicione-os ao estoque pela seção acima para regularizar o cadastro."
                    )
                with st.expander(
                    "Ver equipamentos não encontrados",
                    expanded=False,
                ):
                    st.json(info["not_found"])

            if "Faturar" not in df_period.columns:
                df_period.insert(0, "Faturar", True)

            edited_df = st.data_editor(
                df_period,
                column_config={
                    "Faturar": st.column_config.CheckboxColumn(
                        "Faturar?",
                        default=True,
                    ),
                    "Cliente": st.column_config.TextColumn(disabled=True),
                    "Terminal": st.column_config.TextColumn(disabled=True),
                    "Nº Equipamento": st.column_config.TextColumn(
                        disabled=True
                    ),
                    "Modelo": st.column_config.TextColumn(disabled=True),
                    "Tipo": st.column_config.TextColumn(disabled=True),
                    "Categoria": st.column_config.TextColumn(
                        disabled=True
                    ),
                    "Dias a Faturar": st.column_config.NumberColumn(
                        disabled=True
                    ),
                    "Valor Unitario": st.column_config.NumberColumn(
                        format="R$ %.2f",
                        disabled=True,
                    ),
                    "Valor a Faturar": st.column_config.NumberColumn(
                        format="R$ %.2f",
                        disabled=True,
                    ),
                },
                hide_index=True,
                use_container_width=True,
                height=430,
                key=f"editor_lote_{key_suffix}",
            )

            df_approved = edited_df[
                edited_df["Faturar"] == True
            ].copy()
            approved_by_period[periodo] = df_approved

            totals = build_totals(df_approved)

            c1, c2, c3 = st.columns(3)
            c1.metric(
                "Clientes",
                int(df_approved["Cliente"].nunique()),
            )
            c2.metric(
                "Terminais faturados",
                int(len(df_approved)),
            )
            c3.metric(
                "Valor total",
                _money_br(totals["geral"]),
            )

            if st.checkbox(
                "Exibir resumo por cliente",
                key=f"show_client_summary_{key_suffix}",
                value=False,
            ):
                client_summary = []
                for cliente, df_client in df_approved.groupby("Cliente"):
                    client_totals = build_totals(df_client)
                    client_summary.append(
                        {
                            "Cliente": cliente,
                            "Cheio": client_totals["terminais_cheio"],
                            "Proporcional": client_totals[
                                "terminais_proporcional"
                            ],
                            "Suspensos": client_totals[
                                "terminais_suspensos"
                            ],
                            "GPRS": client_totals["terminais_gprs"],
                            "Satelitais": client_totals[
                                "terminais_satelitais"
                            ],
                            "Valor Total": client_totals["geral"],
                        }
                    )

                if client_summary:
                    st.dataframe(
                        pd.DataFrame(client_summary).sort_values("Cliente"),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Valor Total": st.column_config.NumberColumn(
                                "Valor Total",
                                format="R$ %.2f",
                            ),
                        },
                    )

            col_save, col_prepare = st.columns(2)

            if col_save.button(
                f"Salvar {periodo}",
                key=f"save_period_{key_suffix}",
                type="primary",
                use_container_width=True,
            ):
                if df_approved.empty:
                    st.warning("Não há registros aprovados neste período.")
                elif salvar_historico_lote(
                    df_approved,
                    periodo,
                ):
                    st.success(f"{periodo} salvo e fechado.")

            export_cache = st.session_state.setdefault(
                "billing_export_cache",
                {},
            )
            approved_signature = _frame_signature(df_approved)
            export_key = f"{batch_signature}:{periodo}"
            prepared_export = export_cache.get(export_key)

            if (
                isinstance(prepared_export, dict)
                and prepared_export.get("signature") != approved_signature
            ):
                export_cache.pop(export_key, None)
                prepared_export = None

            if col_prepare.button(
                "Preparar Excel e PDFs",
                key=f"prepare_exports_{key_suffix}",
                use_container_width=True,
            ):
                with st.spinner(
                    f"Gerando arquivos de {periodo} somente uma vez..."
                ):
                    prepared_export = {
                        "signature": approved_signature,
                        "excel": generate_master_excel(df_approved),
                        "pdf_zip": create_zip_of_pdfs(
                            df_approved,
                            periodo,
                        ),
                    }
                    export_cache[export_key] = prepared_export

            if prepared_export:
                download_excel, download_pdf = st.columns(2)
                download_excel.download_button(
                    "Baixar Excel",
                    prepared_export["excel"],
                    f"Faturamento_Lote_{periodo.replace(' ', '_')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"excel_period_{key_suffix}",
                    use_container_width=True,
                )
                download_pdf.download_button(
                    "Baixar PDFs (ZIP)",
                    prepared_export["pdf_zip"],
                    f"PDFs_{periodo.replace(' ', '_')}.zip",
                    "application/zip",
                    key=f"pdf_period_{key_suffix}",
                    use_container_width=True,
                )
            else:
                st.caption(
                    "Excel e PDFs não são gerados durante a navegação. "
                    "Clique em 'Preparar Excel e PDFs' apenas quando precisar baixar."
                )

    st.markdown("---")
    st.subheader("Carga histórica em massa")

    st.warning(
        "Ao salvar todos os períodos, cada mês será gravado separadamente no histórico "
        "e receberá seu próprio fechamento mensal. Faturamentos idênticos já existentes "
        "não criam revisão duplicada."
    )

    confirm_bulk = st.checkbox(
        f"Confirmo o salvamento de {len(ordered_periods)} período(s) no histórico.",
        key="confirm_bulk_history",
    )

    if st.button(
        "Salvar todos os períodos processados",
        type="primary",
        use_container_width=True,
        disabled=not confirm_bulk,
        key="save_all_periods",
    ):
        successes = []
        failures = []

        progress = st.progress(0)
        status = st.empty()

        for index, periodo in enumerate(ordered_periods, start=1):
            frame = approved_by_period.get(periodo)

            status.write(
                f"Salvando {periodo} ({index}/{len(ordered_periods)})..."
            )

            try:
                if frame is None or frame.empty:
                    failures.append(
                        {
                            "Período": periodo,
                            "Motivo": "Nenhum registro aprovado.",
                        }
                    )
                elif salvar_historico_lote(frame, periodo):
                    successes.append(periodo)
                else:
                    failures.append(
                        {
                            "Período": periodo,
                            "Motivo": "Falha ao salvar ou fechar o mês.",
                        }
                    )
            except Exception as exc:
                failures.append(
                    {
                        "Período": periodo,
                        "Motivo": str(exc),
                    }
                )

            progress.progress(index / len(ordered_periods))

        status.empty()

        if successes:
            st.success(
                f"{len(successes)} período(s) salvos com sucesso: "
                + ", ".join(successes)
            )

        if failures:
            st.error(
                f"{len(failures)} período(s) apresentaram falha."
            )
            st.dataframe(
                pd.DataFrame(failures),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.balloons()
            st.success(
                "Carga histórica concluída. Os meses já estão disponíveis "
                "para histórico, analytics e churn."
            )

else:
    st.info(
        "Aguardando o carregamento dos relatórios. Você pode selecionar vários "
        "arquivos de meses diferentes no mesmo envio."
    )
